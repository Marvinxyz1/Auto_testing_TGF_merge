import os
import re
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment

# Directory paths
input_dir = '/Users/marvin/Desktop/py/work/Test/test'
output_dir = '/Users/marvin/Desktop/py/work/Test/tgf_output'

# Identify files with similar names and distinguish main and sub
files = os.listdir(input_dir)
file_pairs = []
processed_files = set()

for f1 in files:
    if f1 in processed_files:
        continue
    f1_parts = re.split('[-_]', f1)
    for f2 in files:
        if f1 == f2 or f2 in processed_files:
            continue
        f2_parts = re.split('[-_]', f2)
        if len(f1_parts) > 2 and len(f2_parts) > 2 and f1_parts[1] == f2_parts[1] and f1_parts[2] == f2_parts[2]:
            if 'before' in f1 and 'after' in f2:
                file_pairs.append((f1, f2))
                processed_files.add(f1)
                processed_files.add(f2)
                break
            elif 'after' in f1 and 'before' in f2:
                file_pairs.append((f2, f1))
                processed_files.add(f1)
                processed_files.add(f2)
                break

# Process each pair of files
for main_file, sub_file in file_pairs:
    main_file_path = os.path.join(input_dir, main_file)
    sub_file_path = os.path.join(input_dir, sub_file)

    # Read the Excel files and the 'TestData' sheets
    main_df = pd.read_excel(main_file_path, sheet_name='TestData')
    sub_df = pd.read_excel(sub_file_path, sheet_name='TestData')

    # Identify common and unique columns
    common_columns = main_df.columns.intersection(sub_df.columns)
    unique_columns_sub = sub_df.columns.difference(main_df.columns)

    # Merge the DataFrames
    merged_df = pd.concat([main_df, sub_df[common_columns]], ignore_index=True)

    # Add unique columns from sub_df to merged_df
    for col in unique_columns_sub:
        merged_df[col] = sub_df[col]

    # Remove columns that are completely empty
    merged_df.dropna(axis=1, how='all', inplace=True)

    # Modify dates to have the month as '08-25'
    for col in merged_df.columns:
        if merged_df[col].dtype == 'object':  # Only process string columns
            merged_df[col] = merged_df[col].apply(lambda x: x[:5] + '08-25' if isinstance(x, str) and x.startswith('2024-') else x)

    # Extract parts for the new file name
    parts = re.split('[-_]', main_file)
    output_file_name = f'TGF_{parts[1]}_{parts[2]}.xlsx'
    output_path = os.path.join(output_dir, output_file_name)

    # Save the merged DataFrame to a new Excel file with sheet name 'TestData'
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        merged_df.to_excel(writer, index=False, sheet_name='TestData')

    # Read the 'TestCases' sheets
    main_testcases_df = pd.read_excel(main_file_path, sheet_name='TestCases')
    sub_testcases_df = pd.read_excel(sub_file_path, sheet_name='TestCases')

    # Vertically merge the DataFrames
    merged_testcases_df = pd.concat([main_testcases_df, sub_testcases_df], ignore_index=True)

    # Save the merged 'TestCases' DataFrame to the same Excel file
    with pd.ExcelWriter(output_path, engine='openpyxl', mode='a') as writer:
        merged_testcases_df.to_excel(writer, index=False, sheet_name='TestCases')

    # Load the original workbook and the new workbook
    original_wb = load_workbook(main_file_path)
    original_ws = original_wb['TestData']
    new_wb = load_workbook(output_path)
    new_ws = new_wb['TestData']

    # Define a border style
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Define font, fill, and alignment styles
    yugothic_font = Font(name='游ゴシック')
    header_font = Font(name='游ゴシック', color='FFFFFF', bold=True)
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    left_alignment = Alignment(horizontal='left')

    # Copy the font and style from the original first sheet to the new sheet
    for row in new_ws.iter_rows(min_row=1, max_row=new_ws.max_row, min_col=1, max_col=new_ws.max_column):
        for cell in row:
            original_cell = original_ws[cell.coordinate] if cell.coordinate in original_ws else None
            if original_cell:
                cell.font = original_cell.font.copy(name='游ゴシック')
                cell.alignment = original_cell.alignment.copy(horizontal='left')
                cell.border = original_cell.border.copy()
                cell.fill = original_cell.fill.copy()
            cell.border = thin_border  # Apply the border to all cells
            cell.font = yugothic_font  # Apply 游ゴシック font to all cells
            cell.alignment = left_alignment  # Align all text to the left

    # Apply header styles to the first row
    for cell in new_ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Adjust column widths
    for col in new_ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        new_ws.column_dimensions[column].width = adjusted_width

    # Save the adjusted workbook
    new_wb.save(output_path)

    # Apply styles to the 'TestCases' sheet
    original_testcases_ws = original_wb['TestCases']
    new_testcases_ws = new_wb['TestCases']

    for row in new_testcases_ws.iter_rows(min_row=1, max_row=new_testcases_ws.max_row, min_col=1, max_col=new_testcases_ws.max_column):
        for cell in row:
            original_cell = original_testcases_ws[cell.coordinate] if cell.coordinate in original_testcases_ws else None
            if original_cell:
                cell.font = original_cell.font.copy(name='游ゴシック')
                cell.alignment = original_cell.alignment.copy(horizontal='left')
                cell.border = original_cell.border.copy()
                cell.fill = original_cell.fill.copy()
            cell.border = thin_border  # Apply the border to all cells
            cell.font = yugothic_font  # Apply 游ゴシック font to all cells
            cell.alignment = left_alignment  # Align all text to the left

    # Apply header styles to the first row of 'TestCases'
    header_fill_orange = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    header_font_black = Font(name='游ゴシック', color='000000', bold=True)
    for cell in new_testcases_ws[1]:
        cell.font = header_font_black
        cell.fill = header_fill_orange

    # Adjust column widths for 'TestCases'
    for col in new_testcases_ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        new_testcases_ws.column_dimensions[column].width = adjusted_width

    # Move 'TestCases' to be the first sheet and 'TestData' to be the second sheet
    new_wb.move_sheet(new_wb['TestCases'], offset=-1)

    # Save the final workbook
    new_wb.save(output_path)

# Copy unpaired files to the output directory
for file in files:
    if file not in processed_files:
        file_path = os.path.join(input_dir, file)
        output_path = os.path.join(output_dir, file)
        os.system(f'cp "{file_path}" "{output_path}"')
        print(f"保留未配对文件: {file}")