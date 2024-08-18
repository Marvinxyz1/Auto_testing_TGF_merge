import os
from openpyxl import load_workbook

# 目标文件夹路径
folder_path = '/Users/marvin/Desktop/py/work/Test/tgf_output'

# 获取文件夹中的所有文件
files = os.listdir(folder_path)

# 识别并删除包含第三个sheet页的Excel文件
files_with_third_sheet = []
for file in files:
    if file.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file)
        workbook = load_workbook(file_path, read_only=True)
        if len(workbook.sheetnames) >= 3:
            files_with_third_sheet.append(file)
            os.remove(file_path)  # 删除文件

# 输出包含第三个sheet页的文件
print("已删除包含第三个sheet页的文件:")
for file in files_with_third_sheet:
    print(file)