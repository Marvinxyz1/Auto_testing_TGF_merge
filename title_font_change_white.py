import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import os

# Define font, fill, and alignment styles
yugothic_font = Font(name='游ゴシック')
header_font_black = Font(name='游ゴシック', color='000000', bold=True)
header_font_white = Font(name='游ゴシック', color='FFFFFF', bold=True)
header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
left_alignment = Alignment(horizontal='left')

# 指定目录路径
directory_path = '/Users/marvin/Desktop/py/work/Test/tgf_output'

# 遍历目录中的所有文件
for filename in os.listdir(directory_path):
    if filename.endswith('.xlsx'):
        file_path = os.path.join(directory_path, filename)
        try:
            workbook = openpyxl.load_workbook(file_path)

            # 获取所有工作表
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # 获取第一行的所有单元格
                for cell in sheet[1]:
                    if sheet_name == 'TestCases':
                        # 设置字体颜色为黑色
                        cell.font = header_font_black
                    elif sheet_name == 'TestData':
                        # 设置字体颜色为白色
                        cell.font = header_font_white

            # 保存修改后的Excel文件
            workbook.save(file_path)
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Invalid Excel file: {file_path}")
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")