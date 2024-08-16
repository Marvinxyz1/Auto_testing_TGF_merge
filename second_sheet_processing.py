import os
import pandas as pd
from openpyxl import load_workbook
from common import input_dir, output_dir, get_file_pairs, save_merged_dataframe, apply_styles_to_sheet, modify_dates

file_pairs = get_file_pairs(input_dir)

for main_file, sub_file in file_pairs:
    main_file_path = os.path.join(input_dir, main_file)
    sub_file_path = os.path.join(input_dir, sub_file)

    main_df = pd.read_excel(main_file_path, sheet_name='TestData')
    sub_df = pd.read_excel(sub_file_path, sheet_name='TestData')

    common_columns = main_df.columns.intersection(sub_df.columns)
    unique_columns_sub = sub_df.columns.difference(main_df.columns)

    merged_df = pd.concat([main_df, sub_df[common_columns]], ignore_index=True)

    for col in unique_columns_sub:
        merged_df[col] = sub_df[col]

    merged_df.dropna(axis=1, how='all', inplace=True)

    merged_df = modify_dates(merged_df)

    output_path = save_merged_dataframe(merged_df, main_file, output_dir)

    original_wb = load_workbook(main_file_path)
    original_ws = original_wb['TestData']
    new_wb = load_workbook(output_path)
    new_ws = new_wb['TestData']

    apply_styles_to_sheet(original_ws, new_ws)

    new_wb.save(output_path)