import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment

# Directory paths
input_dir = '/Users/marvin/Desktop/py/work/Test/tgf'
output_dir = '/Users/marvin/Desktop/py/work/Test/tgf_output'

def get_file_pairs(input_dir):
    files = os.listdir(input_dir)
    file_pairs = []
    for f1 in files:
        for f2 in files:
            if f1 != f2 and f1.split('_')[2] == f2.split('_')[2]:
                if 'before' in f1 and 'after' in f2:
                    file_pairs.append((f1, f2))
                elif 'after' in f1 and 'before' in f2:
                    file_pairs.append((f2, f1))
    return file_pairs

def save_merged_dataframe(merged_df, main_file, output_dir):
    parts = main_file.split('_')
    output_file_name = f'TGF_{parts[1]}_{parts[2]}.xlsx'
    output_path = os.path.join(output_dir, output_file_name)

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        merged_df.to_excel(writer, index=False, sheet_name='TestData')

    return output_path

def apply_styles_to_sheet(original_ws, new_ws):
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    yugothic_font = Font(name='游ゴシック')
    header_font = Font(name='游ゴシック', color='FFFFFF')
    header_fill = PatternFill(start_color='000000', end_color='000000', fill_type='solid')
    left_alignment = Alignment(horizontal='left')

    for row in new_ws.iter_rows(min_row=1, max_row=new_ws.max_row, min_col=1, max_col=new_ws.max_column):
        for cell in row:
            original_cell = original_ws[cell.coordinate] if cell.coordinate in original_ws else None
            if original_cell:
                cell.font = original_cell.font.copy(name='游ゴシック')
                cell.alignment = original_cell.alignment.copy(horizontal='left')
                cell.border = original_cell.border.copy()
                cell.fill = original_cell.fill.copy()
            cell.border = thin_border
            cell.font = yugothic_font
            cell.alignment = left_alignment

    for cell in new_ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for col in new_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        new_ws.column_dimensions[column].width = adjusted_width

def modify_dates(merged_df):
    for col in merged_df.columns:
        if merged_df[col].dtype == 'object':
            merged_df[col] = merged_df[col].apply(lambda x: x[:5] + '08-25' if isinstance(x, str) and x.startswith('2024-') else x)
    return merged_df