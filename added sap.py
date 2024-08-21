import os
import pandas as pd
from openpyxl import load_workbook

# 定义目标文件夹路径
folder_path = '/Users/marvin/Desktop/py/work/Test/tgf_output_副本2'

# 获取文件夹中的所有文件
files = os.listdir(folder_path)

# 处理每个Excel文件
for file in files:
    if file.endswith('.xlsx'):
        file_path = os.path.join(folder_path, file)
        
        # 加载工作簿
        workbook = load_workbook(file_path)
        
        # 处理 TestCases sheet页
        if 'TestCases' in workbook.sheetnames:
            df_testcases = pd.read_excel(file_path, sheet_name='TestCases')
            if 'Case_ID' in df_testcases.columns:
                sap_linkage_indices = df_testcases[df_testcases['Case_ID'].str.contains('sap_linkage', na=False)].index
                for i, idx in enumerate(sap_linkage_indices, start=1):
                    df_testcases.at[idx, 'Case_ID'] = f'sap_linkage_{i}'
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_testcases.to_excel(writer, sheet_name='TestCases', index=False)

        # 处理 TestData sheet页
        if 'TestData' in workbook.sheetnames:
            df_testdata = pd.read_excel(file_path, sheet_name='TestData')
            if 'Case_ID' in df_testdata.columns:
                sap_linkage_indices = df_testdata[df_testdata['Case_ID'].str.contains('sap_linkage', na=False)].index
                for i, idx in enumerate(sap_linkage_indices, start=1):
                    df_testdata.at[idx, 'Case_ID'] = f'sap_linkage_{i}'
                with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                    df_testdata.to_excel(writer, sheet_name='TestData', index=False)

print("处理完成。")